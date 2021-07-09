import requests
import openpyxl

urls=['https://www.cnss.com.cn/u/cms/www/indexJson/bdi_week.json?v=1590297685002',
'https://www.cnss.com.cn/u/cms/www/indexJson/bci_week.json?v=1590297624110',
'https://www.cnss.com.cn/u/cms/www/indexJson/bpi_week.json?v=1590298943084',
'https://www.cnss.com.cn/u/cms/www/indexJson/bsi_week.json?v=1590299130351',
'https://www.cnss.com.cn/u/cms/www/indexJson/bhsi_week.json?v=1590299220997',
'https://www.cnss.com.cn/u/cms/www/indexJson/capesize_week.json?v=1590309574560',
'https://www.cnss.com.cn/u/cms/www/indexJson/panamax_week.json?v=1590309789692',
'https://www.cnss.com.cn/u/cms/www/indexJson/supramax_week.json?v=1590309839740',
'https://www.cnss.com.cn/u/cms/www/indexJson/handysize_week.json?v=1590309905741']


file=openpyxl.Workbook()
sheet1=file.create_sheet('BDI',0)
sheet2=file.create_sheet('BCI',1)
sheet3=file.create_sheet('BPI',2)
sheet4=file.create_sheet('BSI',3)
sheet5=file.create_sheet('BHSI',4)
sheet6=file.create_sheet('Capesize',5)
sheet7=file.create_sheet('Panama',6)
sheet8=file.create_sheet('Supermax',7)
sheet9=file.create_sheet('Handysize',8)


for i in range(9):
    res=requests.get(urls[i])
    js_res=res.json()

    for u in range(5):
        try:
            date=js_res[u]['date']
            index=js_res[u]['index']
        except IndexError:
            continue

        file.worksheets[i].append([date,index])

file.save('波罗的海综合指数和租金.xlsx')
file.close()



    

