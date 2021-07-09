
import openpyxl
import bs4
import requests

def newfile():
    global file
    file=openpyxl.Workbook()

#CCFI
#def ccfi():
    url1='https://www.sse.net.cn/index/singleIndex?indexType=ccfi'
    sheet1=file.create_sheet('CCFI',0)
    res_ccfi=requests.get(url1)
    bs_ccfi=bs4.BeautifulSoup(res_ccfi.text,'html.parser')
    ccfi_index=bs_ccfi.find_all('tbody')

    for line in ccfi_index:
        direction=line.find_all('p')
        scores=line.find_all('td')

    names=['中国出口集装箱运价综合指数','CCFI']
    for i in direction:
        name=i.text.strip()
        names.append(name)

    indexes=[]
    for s in scores:
        score=s.text.strip()
        indexes.append(score) 

    ind=0
    file.worksheets[0].append(['航线','上期','本期','涨跌%'])
    for t in range(13):
        file.worksheets[0].append([names[ind*2],indexes[ind*4+5],indexes[ind*4+6],indexes[ind*4+7]])
        ind=ind+1

#def scfi():
    url2='https://www.sse.net.cn/index/singleIndex?indexType=scfi'
    sheet=file.create_sheet('SCFI',1)
    res_scfi=requests.get(url2)
    bs_scfi=bs4.BeautifulSoup(res_scfi.text,'html.parser')
    scfi_index=bs_scfi.find_all('tbody')

    for line in scfi_index:
        direction=line.find_all('p')
        scores=line.find_all('td')

    names=['上海出口集装箱运价指数','SCFI','分航线 ','(Line Service)']
    for i in direction:
        name=i.text.strip()
        names.append(name)

    indexes=[]
    for s in scores:
        score=s.text.strip()
        indexes.append(score) 
    
    file.worksheets[1].append([indexes[0],indexes[1],indexes[2],indexes[3],indexes[4],indexes[5]])
    file.worksheets[1].append(['上海出口集装箱运价指数',indexes[7],indexes[8],indexes[9],indexes[10],indexes[11]])

    ind=1
    for t in range(13):
        file.worksheets[1].append([names[ind*2+2],indexes[ind*6+7],indexes[ind*6+8],indexes[ind*6+9],indexes[ind*6+10],indexes[ind*6+11]])
        ind=ind+1

#def fdi():
    url3='https://www.sse.net.cn/index/singleIndex?indexType=fdi'
    sheet=file.create_sheet('远东干散货',2)
    res_fdi=requests.get(url3)
    bs_fdi=bs4.BeautifulSoup(res_fdi.text,'html.parser')
    fdi_index=bs_fdi.find_all('tbody')

    for line in fdi_index:
        scores=line.find_all('td')

    indexes=[]
    for s in scores:
        score=s.text.strip()
        indexes.append(score) 

    file.worksheets[2].append([indexes[0],indexes[1],indexes[2],indexes[3],indexes[4],indexes[5],indexes[6]])
    ind=1
    for t in range(26):
        file.worksheets[2].append([indexes[ind*7],indexes[ind*7+1],indexes[ind*7+2],indexes[ind*7+3],indexes[ind*7+4],indexes[ind*7+5],indexes[ind*7+6]])
        ind=ind+1

#def savefile():
    file.save('上海航运交易所数据.xlsx')
    file.close()

newfile()
